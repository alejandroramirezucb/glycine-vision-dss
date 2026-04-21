import os
os.environ["TF_USE_LEGACY_KERAS"] = "1"

import numpy as np
from pathlib import Path
from PIL import Image
import openpyxl
from openpyxl.styles import PatternFill, Font
import tensorflow as tf
from config import (
    test_folder, model_m1_path, model_m2_path, batch_size,
    archivo_excel_salida_m1, archivo_excel_salida_m2,
    disease_classes_m2, binary_classes_m1
)

def load_model_with_fallback(model_path):
    model_path_obj = Path(model_path)
    model_path_str = str(model_path)

    if model_path_obj.is_dir():
        h5_candidates = list(model_path_obj.glob("*.h5")) or list(model_path_obj.rglob("*.h5"))
        if h5_candidates:
            model_path_str = str(h5_candidates[0])

    try:
        return tf.keras.models.load_model(model_path_str)
    except Exception as first_error:
        try:
            from tensorflow.keras.layers import DepthwiseConv2D

            class PatchedDepthwiseConv2D(DepthwiseConv2D):
                def __init__(self, **kwargs):
                    kwargs.pop('groups', None)
                    super().__init__(**kwargs)

            return tf.keras.models.load_model(
                model_path_str,
                custom_objects={"DepthwiseConv2D": PatchedDepthwiseConv2D}
            )
        except Exception:
            raise RuntimeError(f"Cannot load model: {model_path_str}. Error: {first_error}")

def load_images_from_folder(folder_path, target_size=(224, 224)):
    images = []
    for img_file in folder_path.rglob("*"):
        if img_file.is_file() and img_file.suffix.lower() in {".jpg", ".jpeg", ".png"}:
            try:
                img = Image.open(img_file).convert("RGB")
                img = img.resize(target_size)
                images.append(np.array(img) / 255.0)
            except:
                pass
    return np.array(images)

def predict_batch(model, images, batch_size=32):
    predictions = []
    for i in range(0, len(images), batch_size):
        batch = images[i:i+batch_size]
        batch_predictions = model.predict(batch, verbose=0)
        predictions.extend(batch_predictions)
    return np.array(predictions)


def write_excel(output_path, class_names, image_counts, accuracies, precisions, recalls, f1_scores):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_letter, header in zip(["A", "B", "C", "D", "E", "F"],
                                   ["Class", "Test Images", "Accuracy %", "Precision", "Recall", "F1-Score"]):
        cell = ws[f"{col_letter}1"]
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font

    for idx, class_name in enumerate(class_names):
        row = idx + 2
        accuracy_pct = round(accuracies[idx] * 100, 2)
        ws[f"A{row}"] = class_name
        ws[f"B{row}"] = image_counts[idx]
        ws[f"C{row}"] = accuracy_pct
        ws[f"D{row}"] = round(precisions[idx], 4)
        ws[f"E{row}"] = round(recalls[idx], 4)
        ws[f"F{row}"] = round(f1_scores[idx], 4)

        fill_color = "00B050" if accuracy_pct >= 90 else "FFEB9C" if accuracy_pct >= 70 else "FF0000"
        ws[f"C{row}"].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    summary_row = len(class_names) + 3
    overall_accuracy = round(np.mean(accuracies) * 100, 2)
    ws[f"A{summary_row}"] = "Overall"
    ws[f"B{summary_row}"] = sum(image_counts)
    ws[f"C{summary_row}"] = overall_accuracy
    ws[f"C{summary_row}"].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    for col in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 15

    wb.save(output_path)

def evaluate_model_on_classes(model, class_names, base_folder):
    all_images = []
    all_true_labels = []
    image_counts = []

    for idx, class_name in enumerate(class_names):
        class_folder = base_folder / class_name
        if class_folder.exists():
            images = load_images_from_folder(class_folder)
            image_counts.append(len(images))
            if len(images) > 0:
                all_images.extend(images)
                all_true_labels.extend([idx] * len(images))
        else:
            image_counts.append(0)

    if not all_images:
        return [], [], [], [], []

    all_images = np.array(all_images)
    all_true_labels = np.array(all_true_labels)
    predictions = predict_batch(model, all_images, batch_size)
    predicted_labels = np.argmax(predictions, axis=1)


    per_class_accuracy, per_class_precision, per_class_recall, per_class_f1 = [], [], [], []

    for idx in range(len(class_names)):
        tp = np.sum((predicted_labels == idx) & (all_true_labels == idx))
        fp = np.sum((predicted_labels == idx) & (all_true_labels != idx))
        fn = np.sum((predicted_labels != idx) & (all_true_labels == idx))
        total_class = np.sum(all_true_labels == idx)

        class_accuracy = tp / total_class if total_class > 0 else 0
        precision = tp / (tp + fp) if (tp + fp) > 0 else 0
        recall = tp / (tp + fn) if (tp + fn) > 0 else 0
        f1 = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0

        per_class_accuracy.append(class_accuracy)
        per_class_precision.append(precision)
        per_class_recall.append(recall)
        per_class_f1.append(f1)

    return image_counts, per_class_accuracy, per_class_precision, per_class_recall, per_class_f1

def evaluate_m1():
    print("\n=== EVALUATING MODEL 1 (BINARY) ===\n")
    model = load_model_with_fallback(model_m1_path)
    base_folder = test_folder / "clasificacion_binaria"

    counts, accuracies, precisions, recalls, f1s = evaluate_model_on_classes(model, binary_classes_m1, base_folder)
    write_excel(archivo_excel_salida_m1, binary_classes_m1, counts, accuracies, precisions, recalls, f1s)
    print(f"M1 results saved to {archivo_excel_salida_m1}")

def evaluate_m2():
    print("\n=== EVALUATING MODEL 2 (PATHOGEN) ===\n")
    model = load_model_with_fallback(model_m2_path)
    base_folder = test_folder / "clasificacion_patogeno"

    counts, accuracies, precisions, recalls, f1s = evaluate_model_on_classes(model, disease_classes_m2, base_folder)
    write_excel(archivo_excel_salida_m2, disease_classes_m2, counts, accuracies, precisions, recalls, f1s)
    print(f"M2 results saved to {archivo_excel_salida_m2}")

def evaluate_models():
    print("=== MODEL EVALUATION ===")
    evaluate_m1()
    evaluate_m2()
    print("\nEvaluation complete!")

if __name__ == "__main__":
    evaluate_models()
